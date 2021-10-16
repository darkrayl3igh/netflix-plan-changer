import sys
import time
import openpyxl
import os
from datetime import date
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import WebDriverException
from openpyxl.styles import PatternFill

if sys.platform == "darwin":
    PATH = "chromedriver"
else:
    PATH = "chromedriver.exe"

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument("--incognito")

path = "C:\\Users\\darkrayl3igh\\Desktop\\Netflix Plan Changer"
os.chdir(path)

customer_data = openpyxl.load_workbook("customer_data.xlsx")
data = customer_data.get_sheet_by_name("Sheet1")

logs = open("logs.txt", "a")

def save_data(username, password, renew_date, processed_bool, active, comments, count):

    comments = comments.replace("\n", "")

    all_good_fill = PatternFill(start_color='88D8B0', end_color='88D8B0', fill_type='solid')
    error_fill = PatternFill(start_color='FFEEAD', end_color='FFEEAD', fill_type='solid')
    wrong_password_fill = PatternFill(start_color='FF6F69', end_color='FF6F69', fill_type='solid')

    data["A" + str(count)] = username
    data["B" + str(count)] = password
    data["C" + str(count)] = renew_date
    data["D" + str(count)] = processed_bool
    data["E" + str(count)] = active
    data["F" + str(count)] = comments

    if("Error" in comments):
        data["F" + str(count)].fill = error_fill
    elif("All Good" in comments):
        data["F" + str(count)].fill = all_good_fill
    elif("Password" in comments or "Plan Cancelled" in comments or "Membership" in comments):
        data["F" + str(count)].fill = wrong_password_fill

    customer_data.save("customer_data_new.xlsx")

    timestamp = date.today().strftime("[%d/%m/%Y]")

    print(timestamp + " - " + comments + ": " + username + " " + password)
    logs.write(timestamp + " - " + comments + ": " + username + " " + password + "\n")

def lazyload(url):
    while(True):
        try:
            driver.get(url)
            break
        except WebDriverException:
            time.sleep(5)
    time.sleep(10)

def convert(text, flag):
    result = ""
    if(flag == 1):
        values = text.split()
        result = result + values[1][:-1] + "/"
        if(values[0] == "January"):
            result = result + "01/"
        elif(values[0] == "February"):
            result = result + "02/"
        elif(values[0] == "March"):
            result = result + "03/"
        elif(values[0] == "April"):
            result = result + "04/"
        elif(values[0] == "May"):
            result = result + "05/"
        elif(values[0] == "June"):
            result = result + "06/"
        elif(values[0] == "July"):
            result = result + "07/"
        elif(values[0] == "August"):
            result = result + "08/"
        elif(values[0] == "September"):
            result = result + "09/"
        elif(values[0] == "October"):
            result = result + "10/"
        elif(values[0] == "November"):
            result = result + "11/"
        elif(values[0] == "December"):
            result = result + "12/"
        result = result + values[2][2:]
    elif(flag == 2):
        values = text.split("/")
        result = values[1] + "/" + values[0] + "/" + values[2] 
    return result

def main():
    try:
        for i in range(124, data.max_row + 1):
            if(data.cell(i, 5).value == True and data.cell(i, 4).value == False and str(data.cell(i, 3).value).strip() != "--"):
                try:
                    driver = webdriver.Chrome(PATH, chrome_options = chrome_options)

                    username = str(data.cell(i, 1).value).strip()
                    password = str(data.cell(i, 2).value).strip()

                    lazyload("https://www.netflix.com/tr-en/login")

                    while(True):
                        try:
                            driver.find_element_by_xpath("//*[@id=\"id_userLoginId\"]").send_keys(username)
                            driver.find_element_by_xpath("//*[@id=\"id_password\"]").send_keys(password)
                            driver.find_element_by_xpath("//*[@id=\"id_password\"]").send_keys(Keys.RETURN)
                            break
                        except NoSuchElementException:
                            lazyload("https://www.netflix.com/tr-en/login")

                    time.sleep(5)
                    
                    try:
                        driver.find_element_by_xpath("/html/body/div[1]/div/div[3]/div/div/div[1]/div/div[2]/b")
                        save_data(username, password, "--", False, False, "Wrong Password", i)
                        driver.close()
                        continue    
                    except NoSuchElementException:
                        pass
                    
                    lazyload("https://www.netflix.com/YourAccount")

                    try:
                        renew_elem_1 = driver.find_element_by_xpath("/html/body/div[1]/div/div/div[2]/div/div/div[2]/div[2]/section/div/div[2]/div/p/span/b[2]")
                        renew_date = convert(renew_elem_1.text, 2)
                    except NoSuchElementException:
                        try:
                            renew_elem_2 = driver.find_element_by_xpath("/html/body/div[1]/div/div/div[3]/div/div/div[2]/div[1]/section/div[2]/div/div/div[1]/div/div[1]/b")
                            renew_date = convert(renew_elem_2.text, 1)
                        except NoSuchElementException:
                            try:
                                cancelled_elem = driver.find_element_by_xpath("/html/body/div[1]/div/div/div[2]/div/div/div[1]/div/div[2]")
                                save_data(username, password, "--", False, False, "Plan Cancelled", i)
                            except NoSuchElementException:
                                try:
                                    paused_elem = driver.find_element_by_xpath("/html/body/div[1]/div/div/div[1]/div/article/section/h2")
                                    if("PAUSED" in paused_elem.text.upper()):
                                        save_data(username, password, "--", False, False, "Membership Paused", i)
                                except NoSuchElementException:
                                    save_data(username, password, "--", False, True, "Unknown Error", i)
                            driver.close()
                            continue

                    plan = ""
                    try:
                        plan_elem = driver.find_element_by_xpath("/html/body/div[1]/div/div/div[2]/div/div/div[2]/div[2]/section/div/div[1]/div[1]/div/b")
                        plan = plan_elem.text 
                    except NoSuchElementException:
                        save_data(username, password, renew_date, False, True, "Error on Your Account Screen", i)
                        driver.close()
                    
                    
                    
                    if(plan.upper() == "PREMIUM"):
                        save_data(data.cell(i, 1).value, data.cell(i, 2).value, data.cell(i, 3).value, data.cell(i, 4).value, data.cell(i, 5).value, data.cell(i, 6).value, i)
                        driver.close()
                    elif(plan.upper() == "BASIC"):
                        lazyload("https://www.netflix.com/ChangePlan")

                        try:
                            premium_elem_1 = driver.find_element_by_xpath("/html/body/div[1]/div/div/div[2]/div/div/ul/li[3]")
                            premium_elem_1.click()
                        except NoSuchElementException:
                            try:
                                premium_elem_2 = driver.find_element_by_xpath("/html/body/div[1]/div/div/div[3]/div/div/ul/li[3]")
                                premium_elem_2.click()
                            except NoSuchElementException:
                                save_data(username, password, renew_date, False, True, "Error on Change Plan Screen", i)
                                driver.close()

                        time.sleep(2)
                        
                        
                        try:
                            continue_elem_1 = driver.find_element_by_xpath("/html/body/div[1]/div/div/div[2]/div/div/div[2]/button[1]")
                            continue_elem_1.click()
                        except NoSuchElementException:
                            try:
                                continue_elem_2 = driver.find_element_by_xpath("/html/body/div[1]/div/div/div[3]/div/div/div[2]/button[1]")
                                continue_elem_2.click()
                            except NoSuchElementException:
                                save_data(username, password, renew_date, False, True, "Error on Change Plan Screen", i)
                                driver.close()
                        
                        time.sleep(2)

                        try:
                            driver.find_element_by_xpath("//*[@id=\"appMountPoint\"]/div/div/div[2]/div/div/div[3]/div/footer/div/button[1]").click()
                        except:
                            save_data(username, password, renew_date, False, True, "Error on Change Plan Screen", i)
                            driver.close()
                        
                        lazyload("https://www.netflix.com/ChangePlan")

                        try:
                            premium_elem_1 = driver.find_element_by_xpath("/html/body/div[1]/div/div/div[2]/div/div/ul/li[1]")
                            premium_elem_1.click()
                        except NoSuchElementException:
                            try:
                                premium_elem_2 = driver.find_element_by_xpath("/html/body/div[1]/div/div/div[3]/div/div/ul/li[1]")
                                premium_elem_2.click()
                            except NoSuchElementException:
                                save_data(username, password, renew_date, False, True, "Error on Change Plan Screen", i)
                                driver.close()

                        time.sleep(2)
                        
                        
                        try:
                            continue_elem_1 = driver.find_element_by_xpath("/html/body/div[1]/div/div/div[2]/div/div/div[2]/button[1]")
                            continue_elem_1.click()
                        except NoSuchElementException:
                            try:
                                continue_elem_2 = driver.find_element_by_xpath("/html/body/div[1]/div/div/div[3]/div/div/div[2]/button[1]")
                                continue_elem_2.click()
                            except NoSuchElementException:
                                save_data(username, password, renew_date, False, True, "Error on Change Plan Screen", i)
                                driver.close()
                        
                        time.sleep(2)

                        try:
                            driver.find_element_by_xpath("//*[@id=\"appMountPoint\"]/div/div/div[2]/div/div/div[3]/div/footer/div/button[1]").click()
                        except:
                            save_data(username, password, renew_date, False, True, "Error on Change Plan Screen", i)
                            driver.close()

                    driver.close()
                except Exception as e:
                    save_data(data.cell(i, 1).value, data.cell(i, 2).value, data.cell(i, 3).value, data.cell(i, 4).value, data.cell(i, 5).value, "Unknown Error - " + str(e), i)
            else:
                save_data(data.cell(i, 1).value, data.cell(i, 2).value, data.cell(i, 3).value, data.cell(i, 4).value, data.cell(i, 5).value, data.cell(i, 6).value, i)
            
    except Exception as e:
        timestamp = date.today().strftime("[%d/%m/%Y]")
        logs.write(timestamp + " - " + "ERROR: " + str(e) + "\n")

    finally:
        logs.close()

if __name__ == "__main__":
    main()